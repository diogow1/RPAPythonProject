import smtplib
import os
from openpyxl import load_workbook
from email.message import EmailMessage

# Configurações
EMAIL_REMETENTE = ""
SENHA = ""

planilha = load_workbook("../planilhas/faturas_exportadas.xlsx")
aba = planilha.active

# Conexão SMTP
smtp = smtplib.SMTP("smtp.gmail.com", 587)
smtp.starttls()
smtp.login(EMAIL_REMETENTE, SENHA)

for i, linha in enumerate(aba.iter_rows(min_row=2, values_only=True), start=2):
    nome, email, telefone, valor, vencimento, status = linha
    if status == "ENVIADO":
        continue

    assunto = "Fatura Pendente"
    corpo = (
        f"Olá, {nome},\n\n"
        f"Segue anexa sua fatura no valor de R$ {valor:.2f}, com vencimento em {vencimento.strftime('%d/%m/%Y')}.\n"
        f"Em caso de dúvidas, estamos à disposição.\n\n"
        f"Atenciosamente,\nEquipe de Cobrança"
    )

    # Anexar o PDF
    pdf_path = f"../pdfs/fatura_{i-1}.pdf"
    if not os.path.exists(pdf_path):
        print(f"PDF não encontrado: {pdf_path}")
        continue

    msg = EmailMessage()
    msg["From"] = EMAIL_REMETENTE
    msg["To"] = email
    msg["Subject"] = assunto
    msg.set_content(corpo)

    with open(pdf_path, "rb") as f:
        msg.add_attachment(f.read(), maintype="application", subtype="pdf", filename=os.path.basename(pdf_path))

    try:
        smtp.send_message(msg)
        print(f"E-mail enviado para {nome} ({email})")
        #aba[f"F{i}"] = "ENVIADO"
    except Exception as e:
        print(f"Erro ao enviar para {nome}: {e}")

# Salvar e fechar
planilha.save("../planilhas/faturas_exportadas.xlsx")
smtp.quit()
print("Todos os e-mails foram enviados.")
