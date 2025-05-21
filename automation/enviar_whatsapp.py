
from openpyxl import load_workbook
import urllib.parse
from time import sleep
import pyautogui
import os
import webbrowser

planilha = load_workbook("../planilhas/faturas_exportadas.xlsx")
aba = planilha.active

webbrowser.open('https://web.whatsapp.com/')
sleep(10)
#input("Pressione ENTER para continuar...")

for i, linha in enumerate(aba.iter_rows(min_row=2, values_only=True), start=2):
    nome, email, telefone, valor, vencimento, status = linha

    if status == "ENVIADO":
        continue

    mensagem = f"Olá {nome}, seu boleto vence no dia {vencimento.strftime('%d/%m/%Y')}. Favor pagar no link ..."
    link = f"https://web.whatsapp.com/send?phone=55{telefone}&text=" + urllib.parse.quote(mensagem)
    
    webbrowser.open(link)
    sleep(10)

    try:
        seta = pyautogui.locateCenterOnScreen('../automation/seta.png', confidence=0.8)
        if seta:
            pyautogui.click(seta[0], seta[1])
            sleep(2)
            pyautogui.hotkey('ctrl', 'w')
            sleep(2)

            print(f"Mensagem enviada para {nome}")
            aba.cell(row=i, column=6).value = "ENVIADO"
        else:
            raise Exception("Imagem da seta não encontrada")

    except Exception as e:
        print(f"Erro ao enviar para {nome}: {e}")
    sleep(5)

planilha.save("../planilhas/faturas_exportadas.xlsx")
print("Todos os WhatsApps foram enviados.")
