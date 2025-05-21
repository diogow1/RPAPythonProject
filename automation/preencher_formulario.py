import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook

driver = webdriver.Chrome()

wb = load_workbook("../planilhas/clientes.xlsx")
ws = wb.active

for row in ws.iter_rows(min_row=2, values_only=True):
    nome, email, telefone, endereco = row

    driver.get("../frontend/cadastro_cliente.php")

    time.sleep(1) 

    driver.find_element(By.NAME, "nome").send_keys(nome)
    driver.find_element(By.NAME, "email").send_keys(email)
    driver.find_element(By.NAME, "telefone").send_keys(telefone)
    driver.find_element(By.NAME, "endereco").send_keys(endereco)
    driver.find_element(By.XPATH, "//input[@type='submit']").click()

    print(f"Cliente '{nome}' cadastrado.")
    time.sleep(1) 

driver.quit()
print("Todos os clientes foram cadastrados.")
